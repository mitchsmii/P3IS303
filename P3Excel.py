import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from statistics import mean, median

def set_Filters(file_path):
    # Load the Excel workbook and sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # Use the active sheet or specify by name like wb['Sheet1']
    
    # Read the grades from column D (assuming data starts from row 2 to skip the header)
    grades = []
    for row in range(2, sheet.max_row + 1):  # Assuming row 1 is the header
        grade = sheet.cell(row=row, column=4).value  # Column D corresponds to column 4
        if grade is not None:
            grades.append(grade)
    
    # Calculate the summary statistics
    highest_grade = max(grades)
    lowest_grade = min(grades)
    mean_grade = mean(grades)
    median_grade = median(grades)
    num_students = len(grades)
    
    # Write the summary statistics to columns F and G
    sheet['F1'] = 'Statistic'
    sheet['G1'] = 'Value'
    
    sheet['F2'] = 'Highest Grade'
    sheet['G2'] = highest_grade
    
    sheet['F3'] = 'Lowest Grade'
    sheet['G3'] = lowest_grade
    
    sheet['F4'] = 'Mean Grade'
    sheet['G4'] = mean_grade
    
    sheet['F5'] = 'Median Grade'
    sheet['G5'] = median_grade
    
    sheet['F6'] = 'Number of Students'
    sheet['G6'] = num_students
    
    # Save changes
    wb.save(file_path)

# STEP 1 : CREATE SHEETS FOR EACH CLASS
input_worksheet = 'Poorly_Organized_Data_1.xlsx'
output_worksheet = 'IS303p3.xlsx'

# Load the input workbook and get its active sheet
input_wb = openpyxl.load_workbook(input_worksheet)
source_sheet = input_wb.active

# Create a new workbook for the output and remove its default sheet
output_wb = openpyxl.Workbook()
output_wb.remove(output_wb.active)

# Get headers (assuming the first row has headers)
headers = [cell.value for cell in next(source_sheet.iter_rows(min_row=1, max_row=1))]

# Set the header name to search for (change if needed)
subject_header = 'Class Name'
if subject_header not in headers:
    raise ValueError(f"Header '{subject_header}' not found. Check your Excel file's headers.")

subject_col_index = headers.index(subject_header)

# Functions for normalization and capitalization
def normalize_subject(subject):
    return str(subject).strip().lower()

def capitalize_subject(subject):
    return str(subject).strip().title()

# Dictionary to map normalized subjects to the actual sheet names in the output workbook
sheet_map = {}

# Iterate over rows, skipping the header row
for row in source_sheet.iter_rows(min_row=2, values_only=True):
    subject = row[subject_col_index]
    norm_subject = normalize_subject(subject)

    # If sheet for this subject doesn't exist, create it with a capitalized title and store the mapping
    if norm_subject not in sheet_map:
        sheet_name = capitalize_subject(subject)
        output_wb.create_sheet(title=sheet_name)
        output_wb[sheet_name].append(headers)
        sheet_map[norm_subject] = sheet_name
        
    # Append the row to the appropriate sheet using the mapping
    output_wb[sheet_map[norm_subject]].append(row)

# Save the output workbook
output_wb.save(output_worksheet)

# Apply filters to the new worksheet
set_Filters(output_worksheet)
