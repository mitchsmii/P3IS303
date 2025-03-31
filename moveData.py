# Justin Maxwell - Carson Hendrix - Prof. Anderson

# DESCRIPTION: Loops through the data in the excel file and moves it to the corresponding sheet
# So it will split the last name, first name, and ID into a list, then appends the grade to the list as well
# Then puts it on the sheet made for that class


# 

from openpyxl import load_workbook

def organizeData(iCount2):
    currSheet = myworkbook["Grades"]
    myworkbook.active = myworkbook["Grades"]
    data = currSheet["B" + str(iCount)].value
    lName, fName, idNum = data.split('_')
    currSheet = myworkbook[currSheet["A" + str(iCount)].value]
    currSheet["A" + str(iCount2)] = lName
    currSheet["B" + str(iCount2)] = fName
    currSheet["C" + str(iCount2)] = idNum
    currSheet = myworkbook["Grades"]
    myworkbook.active = myworkbook["Grades"]

myworkbook = load_workbook("Poorly_Organized_Data_1.xlsx")
currSheet = myworkbook["Grades"]

iCount = 2

iCount2 = 2
myworkbook.create_sheet("Algebra", 1)
while currSheet["A" + str(iCount)].value == "Algebra":
    organizeData(iCount2)
    iCount += 1
    iCount2 += 1

myworkbook.create_sheet("Trigonometry", 2)
iCount2 = 2
while currSheet["A" + str(iCount)].value == "Trigonometry":
    organizeData(iCount2)
    iCount += 1
    iCount2 += 1

myworkbook.create_sheet("Geometry", 3)
iCount2 = 2
while currSheet["A" + str(iCount)].value == "Geometry":
    organizeData(iCount2)
    iCount += 1
    iCount2 += 1

myworkbook.create_sheet("Calculus", 4)
iCount2 = 2
while currSheet["A" + str(iCount)].value == "Calculus":
    organizeData(iCount2)
    iCount += 1
    iCount2 += 1

myworkbook.create_sheet("Statistics", 5)
iCount2 = 2
while currSheet["A" + str(iCount)].value == "Statistics":
    organizeData(iCount2)
    iCount += 1
    iCount2 += 1

myworkbook.save(filename= "Poorly_Organized_Data_1.xlsx")