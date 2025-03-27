# Will Knudson
# Adding Filters and Summary Information**:
# A filter is added to the columns for easier data manipulation.
# Summary statistics (highest grade, lowest grade, mean grade, median grade, and the number of students)
# are added in columns F and G.

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

from statistics import mean, median

def set_Filters(file_path):
    # Load the Excel workbook and sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active  # Use the active sheet or specify by name like wb['Sheet1']
    
    # Read the grades from column D (assuming data starts from row 2 to skip the header)
    grades = []
    for row in range(2, sheet.max_row + 1):  # Assuming row 1 is the header
        grade = sheet.cell(row=row, column=4).value  # Column D corresponds to column 4
        if grade is not None:
            grades.append(grade)
    
    # Calculate the summary statistics
    highest_grade = max(grades)
    lowest_grade = min(grades)
    mean_grade = mean(grades)
    median_grade = median(grades)
    num_students = len(grades)
    
    # Write the summary statistics to columns F and G
    sheet['F1'] = 'Statistic'
    sheet['G1'] = 'Value'
    
    sheet['F2'] = 'Highest Grade'
    sheet['G2'] = highest_grade
    
    sheet['F3'] = 'Lowest Grade'
    sheet['G3'] = lowest_grade
    
    sheet['F4'] = 'Mean Grade'
    sheet['G4'] = mean_grade
    
    sheet['F5'] = 'Median Grade'
    sheet['G5'] = median_grade
    
    sheet['F6'] = 'Number of Students'
    sheet['G6'] = num_students