# Carson Hendrix, Cohen Scott, Joseph Xiong, Justin Maxwell, Mitchell Smith, Will Knudson
# Section 004
# Professor Anderson
# 04/02/2025

# DESCRIPTION: Loops through the data in the excel file and moves it to the corresponding sheet
# So it will split the last name, first name, and ID into a list, then appends the grade to the list as well
# Then puts it on the sheet made for that class


import openpyxl
from openpyxl import Workbook
from statistics import mean, median
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Modified function to split the data and also transfer the grade.
def organize_data(wb, input_row, output_row):
    grades_sheet = wb["Grades"]
    # Get the raw data from column B of the current row.
    data = grades_sheet["B" + str(input_row)].value
    if data is None:
        return  # Skip if there's no data.
    # Split the data into last name, first name, and id number.
    lName, fName, idNum = data.split('_')
    # Also get the grade from (for example) column C.
    grade = grades_sheet.cell(row=input_row, column=3).value
    # Determine the target sheet by reading the class name from column A.
    target_sheet_name = grades_sheet["A" + str(input_row)].value
    target_sheet = wb[target_sheet_name]
    # Write the split data into the target sheet, including the grade.
    target_sheet["A" + str(output_row)] = lName
    target_sheet["B" + str(output_row)] = fName
    target_sheet["C" + str(output_row)] = idNum
    target_sheet["D" + str(output_row)] = grade

# Function to compute and write summary statistics on each sheet.
def set_filters_all_sheets(file_path, grade_column=4):
    wb = openpyxl.load_workbook(file_path)
    # Loop over each sheet except the "Grades" sheet.
    for sheet_name in wb.sheetnames:
        if sheet_name == "Grades":
            continue  # Skip the Grades sheet if desired.
        sheet = wb[sheet_name]
        grades = []
        # Assume data starts at row 2 (row 1 is header); grade data is in column D.
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=grade_column).value
            if cell_value is not None:
                try:
                    grade = float(cell_value)
                    grades.append(grade)
                except ValueError:
                    print(f"Warning: Could not convert grade '{cell_value}' in sheet {sheet_name} row {row} to a number.")
        if not grades:
            print(f"No grade data found in sheet {sheet_name}. Skipping filtering.")
            continue
        # Calculate summary statistics.
        highest_grade = max(grades)
        lowest_grade = min(grades)
        mean_grade = mean(grades)
        median_grade = median(grades)
        num_students = len(grades)
        # Write summary statistics to columns F and G.
        sheet['F1'] = 'Statistic'
        sheet['G1'] = 'Value'
        sheet['F2'] = 'Highest Grade'
        sheet['G2'] = highest_grade
        sheet['F3'] = 'Lowest Grade'
        sheet['G3'] = lowest_grade
        sheet['F4'] = 'Mean Grade'
        sheet['G4'] = mean_grade
        sheet['F5'] = 'Median Grade'
        sheet['G5'] = median_grade
        sheet['F6'] = 'Number of Students'
        sheet['G6'] = num_students
    wb.save(file_path)

# Define input and output workbook filenames.
input_worksheet = 'Poorly_Organized_Data_1.xlsx'
output_worksheet = 'formatted_grades.xlsx'

# Load the input workbook and get its active sheet.
input_wb = openpyxl.load_workbook(input_worksheet)
source_sheet = input_wb.active

# Create a new workbook for the output and remove its default sheet.
output_wb = Workbook()
output_wb.remove(output_wb.active)

# Create a "Grades" sheet in the output workbook and copy all data from the input file.
grades_sheet = output_wb.create_sheet(title="Grades")
for row in source_sheet.iter_rows(values_only=True):
    grades_sheet.append(row)

# Get headers from the first row (assuming headers are in the first row).
headers = [cell for cell in next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

# Set the header name to search for. This column will be used to create new sheets.
subject_header = 'Class Name'
if subject_header not in headers:
    raise ValueError(f"Header '{subject_header}' not found. Check your Excel file's headers.")

subject_col_index = headers.index(subject_header)

# Functions for normalization and capitalization.
def normalize_subject(subject):
    return str(subject).strip().lower()

def capitalize_subject(subject):
    return str(subject).strip().title()

# Dictionary to map normalized subjects to the actual sheet names in the output workbook.
sheet_map = {}

# Create a new sheet for each class and add a header row for the split data.
# The headers now include "Grade" (column D).
for row in grades_sheet.iter_rows(min_row=2, values_only=True):
    subject = row[subject_col_index]
    norm_subject = normalize_subject(subject)
    if norm_subject not in sheet_map:
        sheet_name = capitalize_subject(subject)
        output_wb.create_sheet(title=sheet_name)
        # Add header row for the split data.
        output_wb[sheet_name].append(["Last Name", "First Name", "ID Number", "Grade"])
        sheet_map[norm_subject] = sheet_name

# Process each row in the "Grades" sheet (starting at row 2 to skip the header)
# and run the organize_data function to split the data into the appropriate class sheet.
for i in range(2, grades_sheet.max_row + 1):
    # Determine the next available row in the target sheet.
    target_sheet_name = grades_sheet["A" + str(i)].value
    target_sheet = output_wb[target_sheet_name]
    output_row = target_sheet.max_row + 1  # Append to the bottom of the target sheet
    organize_data(output_wb, i, output_row)

# Bolds  and  autofits the headers in each sheet
for sheet in output_wb.sheetnames:
    ws = output_wb[sheet]
    
    for col in [1, 2, 3, 4, 6, 7]:  # Only columns A-D, F-G
        header_cell = ws.cell(row=1, column=col)
        if header_cell.value:  # Check if there's a header to avoid errors
            header_cell.font = Font(bold=True)  # Bold the header
            ws.column_dimensions[get_column_letter(col)].width = (len(header_cell.value) + 2)  # Autofit

# Add filters to each sheet in the workbook
# This section was written with the assistance of AI
for sheet_name in output_wb.sheetnames:
    if sheet_name == "Grades":
        continue  # Skip the "Grades" sheet if filters are not needed there
    ws = output_wb[sheet_name]
    # Apply filter to the range covering columns A to D (Last Name, First Name, ID Number, Grade)
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

# Save the output workbook.
output_wb.save(output_worksheet)

# Apply filters (compute summary stats) on each class sheet.
# Here, we assume that in each class sheet the Grade is in column D (i.e. index 4).
set_filters_all_sheets(output_worksheet, grade_column=4)

# output_wb.save(output_worksheet)

# close workbook
output_wb.close()
