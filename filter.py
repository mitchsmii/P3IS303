import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
file_path = "IS303p3.xlsx"  
wb = load_workbook(file_path)

# List of tabs to apply filters
subject_tabs = ["Algebra", "Trigonometry", "Geometry", "Calculus", "Statistics"]

for sheet_name in subject_tabs:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.auto_filter.ref = ws.dimensions  # Apply filter to the whole sheet

# Save the modified file
output_file = "filtered_grades.xlsx"
wb.save(output_file)

print(f"Filters applied successfully! Saved as {output_file}")