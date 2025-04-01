import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from statistics import mean, median

# Function to split the data from the Grades sheet row and write it into the corresponding class sheet.
def organizeData(wb, input_row, output_row):
    grades_sheet = wb["Grades"]
    # Get the raw data from column B of the current row.
    data = grades_sheet["B" + str(input_row)].value
    if data is None:
        return  # Skip if there's no data.
    # Split the data into last name, first name, and id number.
    lName, fName, idNum = data.split('_')
    # Determine the target sheet by reading the class name from column A.
    target_sheet_name = grades_sheet["A" + str(input_row)].value
    target_sheet = wb[target_sheet_name]
    # Write the split data into the target sheet.
    target_sheet["A" + str(output_row)] = lName
    target_sheet["B" + str(output_row)] = fName
    target_sheet["C" + str(output_row)] = idNum

def set_Filters(file_path):
    # Load the Excel workbook.
    wb = openpyxl.load_workbook(file_path)
    # If you intend to use a specific sheet (for example "Grades"), use:
    # sheet = wb["Grades"]
    # Otherwise, the active sheet will be used:
    sheet = wb.active
    
    # Read the grades from column D (assuming data starts from row 2 to skip the header)
    grades = []
    for row in range(2, sheet.max_row + 1):  # Assuming row 1 is the header
        grade = sheet.cell(row=row, column=3).value  # Column D corresponds to column 4
        if grade is not None:
            grades.append(grade)
    
    # If no grade data is found, skip the filtering to avoid errors.
    if not grades:
        print("No grade data found in column D. Skipping filtering.")
        return
    
    # Calculate the summary statistics
    highest_grade = max(grades)
    lowest_grade = min(grades)
    mean_grade = mean(grades)
    median_grade = median(grades)
    num_students = len(grades)
    
    # Write the summary statistics to columns F and G
    sheet['F1'] = 'Statistic'
    sheet['G1'] = 'Value'
    
    sheet['F2'] = 'Highest Grade'
    sheet['G2'] = highest_grade
    
    sheet['F3'] = 'Lowest Grade'
    sheet['G3'] = lowest_grade
    
    sheet['F4'] = 'Mean Grade'
    sheet['G4'] = mean_grade
    
    sheet['F5'] = 'Median Grade'
    sheet['G5'] = median_grade
    
    sheet['F6'] = 'Number of Students'
    sheet['G6'] = num_students
    
    # Save changes
    wb.save(file_path)

# Define input and output workbook filenames.
input_worksheet = 'Poorly_Organized_Data_1.xlsx'
output_worksheet = 'IS303p3.xlsx'

# Load the input workbook and get its active sheet.
input_wb = openpyxl.load_workbook(input_worksheet)
source_sheet = input_wb.active

# Create a new workbook for the output and remove its default sheet.
output_wb = Workbook()
output_wb.remove(output_wb.active)

# Create a "Grades" sheet in the output workbook and copy all data from the input file.
grades_sheet = output_wb.create_sheet(title="Grades")
for row in source_sheet.iter_rows(values_only=True):
    grades_sheet.append(row)

# Get headers from the first row (assuming headers are in the first row).
headers = [cell for cell in next(source_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

# Set the header name to search for. This column will be used to create new sheets.
subject_header = 'Class Name'
if subject_header not in headers:
    raise ValueError(f"Header '{subject_header}' not found. Check your Excel file's headers.")

subject_col_index = headers.index(subject_header)

# Functions for normalization and capitalization.
def normalize_subject(subject):
    return str(subject).strip().lower()

def capitalize_subject(subject):
    return str(subject).strip().title()

# Dictionary to map normalized subjects to the actual sheet names in the output workbook.
sheet_map = {}

# Create a new sheet for each class and add a header row for the split data.
# We assume that the target sheets will contain "Last Name", "First Name", and "ID Number".
for row in grades_sheet.iter_rows(min_row=2, values_only=True):
    subject = row[subject_col_index]
    norm_subject = normalize_subject(subject)
    if norm_subject not in sheet_map:
        sheet_name = capitalize_subject(subject)
        output_wb.create_sheet(title=sheet_name)
        # Add header row for the split data.
        output_wb[sheet_name].append(["Last Name", "First Name", "ID Number"])
        sheet_map[norm_subject] = sheet_name

# Process each row in the "Grades" sheet (starting at row 2 to skip the header)
# and run the organizeData function to split the data into the appropriate class sheet.
for i in range(2, grades_sheet.max_row + 1):
    # Determine the next available row in the target sheet.
    target_sheet_name = grades_sheet["A" + str(i)].value
    target_sheet = output_wb[target_sheet_name]
    output_row = target_sheet.max_row + 1  # Append to the bottom of the target sheet
    organizeData(output_wb, i, output_row)

# Save the output workbook.
output_wb.save(output_worksheet)

# Apply filters (this will now check for empty grade data and skip if none is found)
set_Filters(output_worksheet)
